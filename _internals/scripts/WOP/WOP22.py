# WOP22.py - Workflow Operations Processor
# Refactored for New User Package (Configuration-driven)
# Based on WOP18 with auto-fill defaults on newly created rows
#
# New in WOP22 (refactored): All paths and settings come from config.py
# Users only edit config.py - no hardcoding needed!
#
# New in WOP18: Auto-fill defaults on each newly created row:
#   I (9):  today's date (short)  -> MM/DD/YYYY
#   K (11): "N"
#   W (23): "Need Permit - Need Fixture"
#   X (24): "NPNF"
#   Y (25): "Not Released"
#   AC (29): "N"

import os
import re
import sys
import json
import traceback
from typing import Optional, Tuple, Set
from datetime import date
from pathlib import Path

# Import config from parent directory
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "config"))
import config

import pythoncom
import win32com.client as win32

# deps for the self-contained tasking sheet feature
import tempfile
from openpyxl import load_workbook


def claim_exists_in_table(excel_wb, sheet_name, table_name, claim_number, claim_col_index) -> bool:
    """Return True if claim_number is already present in the table's claim column.
    Uses Excel's COUNTIF so text/number storage both match."""
    try:
        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
    except Exception:
        return False
    try:
        col_rng = tbl.ListColumns(claim_col_index).DataBodyRange
    except Exception:
        col_rng = None
    if not col_rng:
        return False

    app = excel_wb.Application
    wf = app.WorksheetFunction

    val = str(claim_number).strip()
    candidates = [val]
    if re.fullmatch(r"\d+", val):
        try:
            candidates.append(int(val))
        except Exception:
            pass

    for c in candidates:
        try:
            if int(wf.CountIf(col_rng, c)) > 0:
                return True
        except Exception:
            continue
    return False


# ====== CONFIG (now from config.py) ======
# Load all settings from configuration file
MAILBOX_NAME = config.OUTLOOK_SETTINGS["mailbox_name"]
CLAIMS_PATH = config.OUTLOOK_SETTINGS["target_folder_path"]

EXCEL_PATH = config.WOP22_SETTINGS["excel_path"]
SHEET_NAME = config.WOP22_SETTINGS["sheet_name"]
TABLE_NAME = config.WOP22_SETTINGS["table_name"]
TABLE_CLAIM_COL_INDEX = config.WOP22_SETTINGS["table_claim_col_index"]

# District mapping file
DISTRICT_XLSX = config.WOP22_SETTINGS["district_xlsx"]

# Safety toggles
DRY_RUN = config.WOP22_SETTINGS["dry_run"]
LIMIT_N = config.WOP22_SETTINGS["limit_n"]
MODEL_NAME = config.WOP22_SETTINGS["model_name"]
TIMEOUT_SECONDS = config.WOP22_SETTINGS["timeout_seconds"]

# Force-leave flags (unchanged)
ALWAYS_LEAVE_WORKBOOK_OPEN = config.WOP22_SETTINGS["always_leave_workbook_open"]
ALWAYS_LEAVE_EXCEL_RUNNING = config.WOP22_SETTINGS["always_leave_excel_running"]

# === Source selection for MEDIUM/LOW ===
MED_LOW_SOURCE_MODE = config.WOP22_SETTINGS["medium_low_source_mode"]


# ====== Claim Extraction (unchanged) ======

PROMPT_SYSTEM = """You are an AI assistant extracting claim numbers from Outlook email content.

Extract only the single, most recent claim number. A valid claim number must satisfy exactly one of the following:

1. Starts with "904" and is exactly 9 digits long. Do not include any spaces or other charactersâ€”return exactly (e.g.) "904123456"
2. Otherwise, it begins with "TD" followed immediately by digits (any length and not starting with 904). In that case include the full "TDâ€¦" string (e.g. "TD2331627").

â—â— Example claim number edge case â€“ If the claim appears as "TD: 904835662" or "TD 904835662," strip the "TD" and return only "904835662" (with no spaces). 904 Claim number should always be 9 digits.

ðŸš« Do NOT extract addresses, structure numbers, etc.
ðŸš« Only return one claim number, no spaces.

Format:
{
  "Claim Number": ""
}
"""


def _build_full_text_from_mail(mail) -> str:
    try:
        sent_on = str(getattr(mail, "SentOn", "") or "")
    except Exception:
        sent_on = ""
    lines = [
        f"From: {getattr(mail, 'SenderName', '')} <{getattr(getattr(mail, 'Sender', None), 'Address', '')}>",
        f"To: {getattr(mail, 'To', '')}",
        f"CC: {getattr(mail, 'CC', '')}",
        f"Sent: {sent_on}",
        f"Subject: {getattr(mail, 'Subject', '')}",
        "",
        f"{getattr(mail, 'Body', '')}",
    ]
    return "\n".join(lines)


def _sanitize_claim_number(raw: str) -> Optional[str]:
    if not raw:
        return None
    s = str(raw).strip().upper()

    m = re.search(r"\bTD[:\s]+(904\d{6}[A-E]?)\b", s)
    if m:
        s = m.group(1)

    m904 = re.fullmatch(r"(904\d{6})([A-E]?)", s)
    if m904:
        return (m904.group(1) + m904.group(2)).strip()

    mtd = re.fullmatch(r"TD(?!904)\d+", s)
    if mtd:
        return s

    return None


def extract_claim_number_from_email(mail) -> Optional[str]:
    try:
        from dotenv import load_dotenv  # type: ignore
        load_dotenv(os.path.join(config.BASE_DIR, "_internals", ".env"))
    except Exception:
        pass

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("Error: OPENAI_API_KEY not set. Skipping this email.")
        return None

    full_text = _build_full_text_from_mail(mail)

    try:
        from openai import OpenAI  # type: ignore
        client = OpenAI(api_key=api_key)
        user_msg = f"Email:\n{full_text.strip()}"
        resp = client.chat.completions.create(
            model=MODEL_NAME,
            temperature=0,
            timeout=TIMEOUT_SECONDS,
            messages=[
                {"role": "system", "content": PROMPT_SYSTEM},
                {"role": "user", "content": user_msg},
            ],
        )
        content = resp.choices[0].message.content or ""
        claim_value = None
        try:
            data = json.loads(content)
            claim_value = data.get("Claim Number")
        except Exception:
            m = re.search(r'"Claim Number"\s*:\s*"([^"]+)"', content)
            if m:
                claim_value = m.group(1)

        return _sanitize_claim_number(claim_value or "")
    except Exception as e:
        print(f"Error: OpenAI extraction failed â€” {e}")
        return None


# ====== Attachment filename scan (unchanged) ======

PAT_904 = re.compile(r"\b904\d{6}[A-E]?\b", re.IGNORECASE)
PAT_TD  = re.compile(r"\bTD(?!904)\d+\b", re.IGNORECASE)

def claims_from_attachment_filenames(mail) -> Set[str]:
    claims: Set[str] = set()
    try:
        atts = getattr(mail, "Attachments", None)
        if not atts:
            return claims
        count = int(getattr(atts, "Count", 0))
        for i in range(1, count + 1):
            try:
                att = atts.Item(i)
                name = str(getattr(att, "FileName", "") or "")
                for m in PAT_904.finditer(name):
                    claims.add(m.group(0).upper())
                for m in PAT_TD.finditer(name):
                    claims.add(m.group(0).upper())
            except Exception:
                continue
    except Exception:
        pass
    return claims


# ====== Excel (COM) (unchanged add) ======

def open_excel_workbook(excel, path) -> Tuple[object, bool]:
    for wb in excel.Workbooks:
        try:
            if os.path.abspath(wb.FullName).lower() == os.path.abspath(path).lower():
                return wb, False
        except Exception:
            continue
    return excel.Workbooks.Open(path), True


def append_claim_to_table(excel_wb, sheet_name, table_name, claim_number, claim_col_index, highlight=False):
    ws = excel_wb.Worksheets(sheet_name)
    tbl = ws.ListObjects(table_name)
    new_row = tbl.ListRows.Add()

    target_cell = new_row.Range.Cells(1, claim_col_index)
    try:
        target_cell.NumberFormat = "General"
    except Exception:
        pass

    val = str(claim_number)
    if re.fullmatch(r"\d+", val):
        try:
            target_cell.Value = int(val)
        except Exception:
            target_cell.Value = val
    else:
        target_cell.Value = val

    if highlight:
        try:
            target_cell.Interior.ColorIndex = 6
        except Exception:
            pass


# ====== Confidence + Prompt (unchanged) ======

def confidence_level(email_claim: Optional[str], tasking_claims: Set[str]) -> str:
    if email_claim and email_claim in tasking_claims:
        return "High"
    if tasking_claims and (not email_claim or email_claim not in tasking_claims):
        return "Low"
    return "Medium"


def prompt_medium_low() -> str:
    """Headless: auto-accept."""
    return "1"
# ====== Self-contained tasking sheet LLM feature (unchanged) ======

def process_tasking_sheet_via_llm(mail) -> Optional[str]:
    try:
        atts = getattr(mail, "Attachments", None)
        if not atts or int(getattr(atts, "Count", 0)) < 1:
            msg = "(no tasking sheet parsed)"
            print("\n" + msg + "\n" + "-"*60 + "\n")
            return None

        # Find first .xlsx attachment
        x_att = None
        x_name = None
        count = int(getattr(atts, "Count", 0))
        for i in range(1, count + 1):
            att = atts.Item(i)
            name = str(getattr(att, "FileName", "") or "")
            if name.lower().endswith(".xlsx"):
                x_att = att
                x_name = name
                break

        if x_att is None or not x_name:
            msg = "(no tasking sheet parsed)"
            print("\n" + msg + "\n" + "-"*60 + "\n")
            return None

        # Save to temp
        tmpdir = tempfile.mkdtemp(prefix="wop12_")
        save_path = os.path.join(tmpdir, x_name)
        x_att.SaveAsFile(save_path)

        # Read B1:B7
        wb = load_workbook(save_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            lines = []
            for r in range(1, 8):
                v = ws.cell(row=r, column=2).value  # Column B
                if v is not None:
                    lines.append(str(v).strip())
            parsed_block = "\n".join(lines).strip()
        finally:
            wb.close()

        if not parsed_block:
            msg = "(no tasking sheet parsed)"
            print("\n" + msg + "\n" + "-"*60 + "\n")
            return None

        prompt_text = f"""You are reading parsed data from and excel spreadsheet. Here is the parsed section of info you need to process...

Start of Data...
{parsed_block}
Data End...

Please determine the following items from parsed data. If no value can be determined, write "none".

"Claim Number, Address, Structure Number, Permit Status"

Please format you answer like this.
Claim# - [determined claim number]
Address - [determined address]
Structure# - [determined structure#]
Permit Status - [determined permit status]

END PROMPT***"""

        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            msg = "(no tasking sheet parsed)"
            print("\n" + msg + "\n" + "-"*60 + "\n")
            return None

        from openai import OpenAI  # imported locally to keep feature self-contained
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model=MODEL_NAME,
            temperature=0,
            timeout=TIMEOUT_SECONDS,
            messages=[{"role": "user", "content": prompt_text}],
        )
        content = (resp.choices[0].message.content or "").strip()
        if content:
            print("\n" + content + "\n" + "-"*60 + "\n")
            return content
        else:
            msg = "(no tasking sheet parsed)"
            print("\n" + msg + "\n" + "-"*60 + "\n")
            return None

    except Exception:
        print("\n(no tasking sheet parsed)\n" + "-"*60 + "\n")
        return None
    finally:
        try:
            if 'save_path' in locals() and os.path.exists(save_path):
                try:
                    os.remove(save_path)
                except Exception:
                    pass
            if 'tmpdir' in locals() and os.path.isdir(tmpdir):
                try:
                    os.rmdir(tmpdir)
                except Exception:
                    pass
        except Exception:
            pass


# ====== Parse LLM output (unchanged) ======

def parse_tasking_llm_output(text: str) -> Optional[dict]:
    if not text:
        return None
    addr_m = re.search(r'(?im)^\s*Address\s*-\s*(.+?)\s*$', text)
    struct_m = re.search(r'(?im)^\s*Structure\s*#?\s*-\s*(.+?)\s*$', text)
    permit_m = re.search(r'(?im)^\s*Permit\s*Status\s*-\s*(.+?)\s*$', text)
    if not (addr_m or struct_m or permit_m):
        return None
    return {
        "address": addr_m.group(1).strip() if addr_m else None,
        "structure": struct_m.group(1).strip() if struct_m else None,
        "permit": permit_m.group(1).strip() if permit_m else None,
    }


# ====== Structure normalizer (unchanged from WOP17) ======

def normalize_structure_number(text: Optional[str]) -> Optional[str]:
    if text is None:
        return None
    s = str(text).upper().strip()
    m = re.search(r'(\d{5,10})', s)
    if not m:
        return s
    digits = m.group(1)
    return digits + "E"

# NEW (fixed): Apply "N - <Permit Status>" to Column U (21).
#              Col K (11) stays "N" from fill_default_fields_on_last_row.
def apply_permit_tag_to_colK(excel_wb, sheet_name, table_name) -> None:
    """
    Reads Permit Status from Column U (21) on the last row.
    - If non-blank -> writes "N - <Permit Status>" back to Column U (21)
    - If blank     -> writes "N" to Column U (21)
    Col K (11) is not modified here (remains "N" from defaults).
    """
    try:
        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
        last_idx = int(tbl.ListRows.Count)
        lr = tbl.ListRows(last_idx).Range

        raw_permit = lr.Cells(1, 21).Value  # Column U (current LLM output, may be blank)
        permit = (str(raw_permit).strip() if raw_permit is not None else "")

        lr.Cells(1, 21).Value = f"N - {permit}" if permit else "N"  # write back to U
        # DO NOT touch K here; K is already "N" from fill_default_fields_on_last_row
    except Exception:
        pass


# ====== NEW (WOP18): Auto-fill defaults on the last row ======

def fill_default_fields_on_last_row(excel_wb, sheet_name, table_name) -> None:
    """
    Writes default values into the last row of MasterTable:
      I(9)=today (short date), K(11)="N", W(23)="Need Permit - Need Fixture",
      X(24)="NPNF", Y(25)="Not Released", AC(29)="N"
    """
    try:
        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
        last_idx = int(tbl.ListRows.Count)
        lr = tbl.ListRows(last_idx).Range

        # Col I (9): today's date (Excel short date)
        try:
            cell_I = lr.Cells(1, 9)
            cell_I.NumberFormat = "mm/dd/yyyy"
            cell_I.Value = date.today().strftime("%m/%d/%Y")
        except Exception:
            pass

        # Col K (11): "N"
        try:
            lr.Cells(1, 11).Value = "N"
        except Exception:
            pass

        # Col W (23): "Need Permit - Need Fixture"
        try:
            lr.Cells(1, 23).Value = "Need Permit - Need Fixture"
        except Exception:
            pass

        # Col X (24): "NPNF"
        try:
            lr.Cells(1, 24).Value = "NPNF"
        except Exception:
            pass

        # Col Y (25): "Not Released"
        try:
            lr.Cells(1, 25).Value = "Not Released"
        except Exception:
            pass

        # Col AC (29): "N"
        try:
            lr.Cells(1, 29).Value = "N"
        except Exception:
            pass

    except Exception:
        # Silent no-op per your minimal-change rule
        pass

# Treat placeholder/unknown permit text as blank
def _clean_permit_status(text):
    """
    Returns a cleaned permit string, or None if it's effectively blank.
    Blanks out things like: [determined permit status], [anything], none, n/a, unknown.
    """
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None

    s_low = s.lower()

    # common "no value" tokens
    if s_low in {"none", "n/a", "na", "n.a.", "unknown"}:
        return None

    # bracketed placeholders (e.g., [determined permit status], [permit], etc.)
    if re.fullmatch(r"\[[^\]]+\]", s) or "determined permit status" in s_low:
        return None

    return s


# ====== Fallback: LLM on raw email body (no cleaning) ======

def process_email_body_via_llm(mail) -> Optional[str]:
    """
    Calls the LLM with the user's exact fallback prompt using the FULL raw email body (no trimming).
    Returns the raw LLM text (may be None on error).
    """
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        print("Error: OPENAI_API_KEY not set. Skipping body fallback.")
        return None

    parsed_block = str(getattr(mail, "Body", "") or "")
    try:
        from openai import OpenAI  # type: ignore
        client = OpenAI(api_key=api_key)
        prompt = (
            'You are reading parsed data from the body of an email. Here is the parsed section of info you need to process... \n\n'
            'Start of Data... \n'
            f'{parsed_block}\n'
            'Data End... \n\n'
            'Please determine the following items from parsed data. If no value can be determined, write "none". \n\n'
            '"Claim Number, Address, Structure Number, Permit Status" \n\n'
            'Please format you answer like this. \n'
            'Claim# - [determined claim number] \n'
            'Address - [determined address] \n'
            'Structure# - [determined structure#] \n'
            'Permit Status - [determined permit status]\n\n'
            'END PROMPT***'
        )
        resp = client.chat.completions.create(
            model=MODEL_NAME,
            temperature=0,
            timeout=TIMEOUT_SECONDS,
            messages=[
                {"role": "user", "content": prompt},
            ],
        )
        return (resp.choices[0].message.content or "").strip()
    except Exception as e:
        print("LLM body fallback error:", e)
        return None


def parse_body_llm_output(text: str) -> Optional[dict]:
    """
    Parse the fallback LLM output. Reuses parse_tasking_llm_output for address/structure/permit,
    and adds claim extraction from the 'Claim# - ...' line.
    """
    if not text:
        return None
    out = parse_tasking_llm_output(text) or {}
    m_claim = re.search(r'(?im)^\s*Claim#\s*-\s*(.+?)\s*$', text)
    if m_claim:
        claim_raw = m_claim.group(1).strip()
        out["claim"] = _sanitize_claim_number(claim_raw) or claim_raw
    return out or None


def _is_noneish(val) -> bool:
    if val is None:
        return True
    s = str(val).strip().lower()
    if not s:
        return True
    if s in {"none", "n/a", "na", "n.a.", "unknown"}:
        return True
    if re.fullmatch(r"\[[^\]]+\]", s):
        return True
    return False

def fallback_from_body_always(excel_wb, sheet_name, table_name, mail, tasking_fields: Optional[dict]) -> None:
    """
    Always runs the email-body LLM parser, even if tasking sheet fields exist.
    - Fills missing structure/address/permit if blank.
    - Logs differences if both tasking and email provide different values.
    """
    fb_text = process_email_body_via_llm(mail)
    fb_fields = parse_body_llm_output(fb_text) if fb_text else None
    if not fb_fields:
        print("(fallback) LLM found no values in email body.")
        return

    print("\nðŸ“¬ Email fallback fields:")
    print(f"  Claim#:   {fb_fields.get('claim')}")
    print(f"  Address:  {fb_fields.get('address')}")
    print(f"  Structure:{fb_fields.get('structure')}")
    print(f"  Permit:   {fb_fields.get('permit')}")

    try:
        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
        last_idx = int(tbl.ListRows.Count)
        lr = tbl.ListRows(last_idx).Range

        # Structure (col E)
        try:
            current_struct = lr.Cells(1, 5).Value
        except Exception:
            current_struct = None
        new_struct = normalize_structure_number(fb_fields.get("structure"))
        if _is_noneish(current_struct) and not _is_noneish(new_struct):
            lr.Cells(1, 5).Value = new_struct
        elif current_struct and new_struct and str(current_struct).strip() != str(new_struct).strip():
            print(f"âš ï¸ Structure differs: Sheet='{current_struct}' vs Email='{new_struct}'")

        # Address (col F)
        try:
            current_addr = lr.Cells(1, 6).Value
        except Exception:
            current_addr = None
        new_addr = fb_fields.get("address")
        if _is_noneish(current_addr) and not _is_noneish(new_addr):
            lr.Cells(1, 6).Value = new_addr
        elif current_addr and new_addr and str(current_addr).strip() != str(new_addr).strip():
            print(f"âš ï¸ Address differs: Sheet='{current_addr}' vs Email='{new_addr}'")

        # Permit (col U)
        try:
            current_permit = lr.Cells(1, 21).Value
        except Exception:
            current_permit = None
        new_permit = _clean_permit_status(fb_fields.get("permit"))
        if _is_noneish(current_permit) and new_permit is not None:
            lr.Cells(1, 21).Value = new_permit
        elif current_permit and new_permit and str(current_permit).strip() != str(new_permit).strip():
            print(f"âš ï¸ Permit differs: Sheet='{current_permit}' vs Email='{new_permit}'")

    except Exception as e:
        print("(fallback) write error:", e)

    print("\n")


# ====== Increment Column A (unchanged) ======

def increment_colA_from_previous_row(excel_wb, sheet_name, table_name) -> None:
    try:
        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
        last_idx = int(tbl.ListRows.Count)
        prev_cell = tbl.ListRows(last_idx - 1).Range.Cells(1, 1)
        prev_val = prev_cell.Value
        new_val = int(prev_val) + 1
        cur_cell = tbl.ListRows(last_idx).Range.Cells(1, 1)
        try:
            cur_cell.NumberFormat = "General"
        except Exception:
            pass
        cur_cell.Value = new_val
    except Exception:
        pass


# ====== District mapping (unchanged) ======

def set_district_from_sender(excel_wb, sheet_name, table_name, district_xlsx_path: str, sender_email: str) -> None:
    try:
        wb_map = load_workbook(district_xlsx_path, read_only=True, data_only=True)
        try:
            ws_map = wb_map.worksheets[0]
            mapping = {}
            max_row = ws_map.max_row or 0
            for r in range(1, max_row + 1):
                district = ws_map.cell(row=r, column=1).value
                if district is None:
                    continue
                district = str(district).strip()
                for c in range(2, 12):  # B..K (10 columns for emails)
                    ev = ws_map.cell(row=r, column=c).value
                    if ev is None:
                        continue
                    email_key = str(ev).strip().lower()
                    if email_key:
                        mapping[email_key] = district
        finally:
            wb_map.close()

        sender_key = (sender_email or "").strip().lower()
        out_val = mapping.get(sender_key, sender_email)

        ws = excel_wb.Worksheets(sheet_name)
        tbl = ws.ListObjects(table_name)
        last_idx = int(tbl.ListRows.Count)
        lr = tbl.ListRows(last_idx).Range
        lr.Cells(1, 3).Value = out_val  # Column C
    except Exception:
        pass


# ====== SMTP normalization & Emilio detection (unchanged) ======

def _get_smtp_sender(mail) -> str:
    try:
        addr_type = str(getattr(mail, "SenderEmailType", "") or "").upper()
        if addr_type == "SMTP":
            return str(getattr(mail, "SenderEmailAddress", "") or "")
        try:
            ex_user = getattr(mail, "Sender", None)
            if ex_user is not None:
                ex_user = ex_user.GetExchangeUser()
                if ex_user is not None:
                    smtp = getattr(ex_user, "PrimarySmtpAddress", None)
                    if smtp:
                        return str(smtp)
        except Exception:
            pass
        return str(getattr(mail, "SenderEmailAddress", "") or "")
    except Exception:
        return ""


def _sender_is_emilio(mail) -> bool:
    """Check if sender is a specific colleague (customize for your team)"""
    try:
        smtp = _get_smtp_sender(mail).strip().lower()
        if smtp == "colleague@company.com":  # Replace with actual email
            return True
        name = str(getattr(mail, "SenderName", "") or "").strip().lower()
        if name == "colleague name":  # Replace with actual name
            return True
        raw_addr = str(getattr(mail, "SenderEmailAddress", "") or "")
        if re.search(r"colleague", raw_addr, flags=re.IGNORECASE):  # Replace pattern
            return True
        return False
    except Exception:
        return False


def effective_sender_for_district(mail) -> str:
    try:
        if not _sender_is_emilio(mail):
            return _get_smtp_sender(mail)
        body = str(getattr(mail, "Body", "") or "")
        m_line = re.search(r'(?im)^\s*from\s*:\s*(.+)$', body)
        if m_line:
            line_text = m_line.group(1)
            m_brackets = re.search(r'<\s*([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})\s*>', line_text)
            if m_brackets:
                return m_brackets.group(1).strip()
            m_email = re.search(r'([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})', line_text)
            if m_email:
                return m_email.group(1).strip()
        return "colleague@company.com"  # Fallback email - customize as needed
    except Exception:
        return _get_smtp_sender(mail)


# ====== Main ======

import os
import shutil
import win32com.client

def ensure_clean_genpy():
    try:
        import win32com.client.gencache as gencache
        gencache.GetModuleForProgID("Excel.Application")
    except AttributeError as e:
        genpy_path = os.path.join(os.environ["LOCALAPPDATA"], "Temp", "gen_py")
        if os.path.exists(genpy_path):
            print(f"âš ï¸ Detected broken gen_py cache. Deleting: {genpy_path}")
            try:
                shutil.rmtree(genpy_path)
                print("âœ… gen_py cache cleared.")
            except Exception as cleanup_error:
                print(f"âŒ Failed to delete gen_py: {cleanup_error}")
        try:
            win32com.client.gencache.Rebuild()
            print("ðŸ”„ Rebuilt COM cache.")
        except Exception as rebuild_error:
            print(f"âŒ Failed to rebuild COM cache: {rebuild_error}")



def process_claims():
    pythoncom.CoInitialize()

    outlook = None
    excel = None
    wb = None
    wrote = 0
    skipped = 0
    excel_was_running = False
    opened_here = False

    try:
        # Outlook
        ensure_clean_genpy()
        outlook = win32.Dispatch("Outlook.Application")
        session = outlook.GetNamespace("MAPI")
        root = session.Folders.Item(MAILBOX_NAME)
        folder = root
        for name in CLAIMS_PATH:
            folder = folder.Folders.Item(name)

        # Excel (attach or start)
        try:
            ensure_clean_genpy()
            excel = win32.GetActiveObject("Excel.Application")
            excel_was_running = True
        except Exception:
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            excel_was_running = False
        if not excel_was_running:
            excel.Visible = False


        # Headless: avoid save/update popups
        try:
            excel.DisplayAlerts = False
        except Exception:
            pass
        wb, opened_here = open_excel_workbook(excel, EXCEL_PATH)

        # Items newest-first
        items = folder.Items
        items.Sort("[ReceivedTime]", True)
        count = 0

        for mail in items:
            try:
                if int(getattr(mail, "Class", 0)) != 43:
                    continue

                count += 1
                if LIMIT_N and count > LIMIT_N:
                    break

                subject = str(getattr(mail, "Subject", "") or "")
                email_claim = extract_claim_number_from_email(mail)
                tasking_claims = claims_from_attachment_filenames(mail)

                print(f"Email subject: {subject}")
                print(f"Extracted claim number from email = {email_claim if email_claim else 'None'}")
                print(f"Extracted claim numbers from tasking sheet = {sorted(tasking_claims) if tasking_claims else 'None'}")

                conf = confidence_level(email_claim, tasking_claims)
                print(f"confidence = {conf}")

                final_claim = email_claim
                highlight = False

                if conf == "High":
                    if final_claim:
                        # CHECK FOR DUPLICATE FIRST (before expensive LLM calls)
                        if claim_exists_in_table(wb, SHEET_NAME, TABLE_NAME, final_claim, TABLE_CLAIM_COL_INDEX):
                            print(f"âš ï¸  Duplicate claim {final_claim} â€” skipping (no LLM extraction for duplicates).")
                            print("\n")
                            skipped += 1
                            continue

                        # NOT A DUPLICATE - Now extract fields via LLM for visibility
                        text = process_tasking_sheet_via_llm(mail)
                        fields = parse_tasking_llm_output(text) if text else None

                        # ADD TO TABLE FIRST (before fallback comparison)
                        # This ensures fallback comparison reads the correct row from Excel
                        append_claim_to_table(wb, SHEET_NAME, TABLE_NAME, final_claim, TABLE_CLAIM_COL_INDEX, highlight=False)
                        wrote += 1

                        # Write tasking sheet extracted fields to E/F/U
                        if fields:
                            ws = wb.Worksheets(SHEET_NAME)
                            tbl = ws.ListObjects(TABLE_NAME)
                            last_idx = int(tbl.ListRows.Count)
                            lr = tbl.ListRows(last_idx).Range
                            try:
                                if fields.get("structure") is not None:
                                    lr.Cells(1, 5).Value = normalize_structure_number(fields.get("structure"))
                            except Exception:
                                pass
                            try:
                                if fields.get("address") is not None:
                                    lr.Cells(1, 6).Value = fields.get("address")
                            except Exception:
                                pass
                            try:
                                permit_val = _clean_permit_status(fields.get("permit"))
                                if permit_val is not None:
                                    lr.Cells(1, 21).Value = permit_val

                            except Exception:
                                pass

                        # NOW compare fallback (after row is added and tasking fields are written)
                        fallback_from_body_always(wb, SHEET_NAME, TABLE_NAME, mail, fields)

                        # Increment A, set District
                        increment_colA_from_previous_row(wb, SHEET_NAME, TABLE_NAME)
                        lookup_email = effective_sender_for_district(mail)
                        set_district_from_sender(wb, SHEET_NAME, TABLE_NAME, DISTRICT_XLSX, lookup_email)

                        # NEW: Always fill defaults for the new row
                        fill_default_fields_on_last_row(wb, SHEET_NAME, TABLE_NAME)

                        apply_permit_tag_to_colK(wb, SHEET_NAME, TABLE_NAME)


                    else:
                        print("Error: High confidence but no email claim parsed. Skipping.\n")
                        skipped += 1
                    continue

                # MEDIUM or LOW â†’ prompt
                # MEDIUM or LOW confidence branch
                choice = prompt_medium_low()
                if choice == 3:
                    print("Skipping this email.\n")
                    skipped += 1
                    continue
                elif choice == 4:
                    print("Edit mode not implemented here.\n")
                    skipped += 1
                    continue

                highlight = (choice == 2)

                append_claim_to_table(
                    wb, SHEET_NAME, TABLE_NAME, final_claim,
                    TABLE_CLAIM_COL_INDEX, highlight=highlight
                )
                wrote += 1

                # === Parse both tasking sheet and email body ===
                tasking_text = process_tasking_sheet_via_llm(mail)
                tasking_fields = parse_tasking_llm_output(tasking_text) if tasking_text else {}

                email_text = process_email_body_via_llm(mail)
                email_fields = parse_body_llm_output(email_text) if email_text else {}

                # === Show side-by-side comparison ===
                print("\nðŸ“Š Parsed Data Comparison (Tasking Sheet vs Email):")
                print("---------------------------------------------------")
                print(f"Claim #:     Sheet = {tasking_fields.get('claim')}   |   Email = {email_fields.get('claim')}")
                print(f"Structure #: Sheet = {tasking_fields.get('structure')}   |   Email = {email_fields.get('structure')}")
                print(f"Address:     Sheet = {tasking_fields.get('address')}   |   Email = {email_fields.get('address')}")
                print(f"Permit:      Sheet = {tasking_fields.get('permit')}   |   Email = {email_fields.get('permit')}")
                print("---------------------------------------------------\n")

                # === Decide source mode ===
                mode = MED_LOW_SOURCE_MODE
                if mode == "prompt":
                    while True:
                        choice = input("Use [S]heet, [E]mail, or [A]uto (prefer sheet, fallback to email)? ").strip().lower()
                        if choice in {"s", "e", "a"}:
                            mode = {"s": "sheet", "e": "email", "a": "auto"}[choice]
                            break
                        else:
                            print("Invalid choice. Please enter S, E, or A.")

                use_email = (mode == "email")
                use_sheet = (mode == "sheet")
                use_auto  = (mode == "auto")

                ws = wb.Worksheets(SHEET_NAME)
                tbl = ws.ListObjects(TABLE_NAME)
                last_idx = int(tbl.ListRows.Count)
                lr = tbl.ListRows(last_idx).Range

                # STRUCTURE (col E)
                struct_sheet = normalize_structure_number(tasking_fields.get("structure"))
                struct_email = normalize_structure_number(email_fields.get("structure"))
                if use_email and struct_email:
                    lr.Cells(1, 5).Value = struct_email
                elif use_sheet and struct_sheet:
                    lr.Cells(1, 5).Value = struct_sheet
                elif use_auto:
                    lr.Cells(1, 5).Value = struct_sheet or struct_email

                # ADDRESS (col F)
                addr_sheet = tasking_fields.get("address")
                addr_email = email_fields.get("address")
                if use_email and addr_email:
                    lr.Cells(1, 6).Value = addr_email
                elif use_sheet and addr_sheet:
                    lr.Cells(1, 6).Value = addr_sheet
                elif use_auto:
                    lr.Cells(1, 6).Value = addr_sheet or addr_email

                # PERMIT (col U)
                permit_sheet = _clean_permit_status(tasking_fields.get("permit"))
                permit_email = _clean_permit_status(email_fields.get("permit"))
                if use_email and permit_email is not None:
                    lr.Cells(1, 21).Value = permit_email
                elif use_sheet and permit_sheet is not None:
                    lr.Cells(1, 21).Value = permit_sheet
                elif use_auto:
                    lr.Cells(1, 21).Value = permit_sheet or permit_email

                print(f"âœ… Using source mode: {mode.upper()} â€” data written based on that choice.\n")


                    # Increment A, set District
                increment_colA_from_previous_row(wb, SHEET_NAME, TABLE_NAME)
                lookup_email = effective_sender_for_district(mail)
                set_district_from_sender(wb, SHEET_NAME, TABLE_NAME, DISTRICT_XLSX, lookup_email)

                # NEW: Always fill defaults for the new row
                fill_default_fields_on_last_row(wb, SHEET_NAME, TABLE_NAME)

                apply_permit_tag_to_colK(wb, SHEET_NAME, TABLE_NAME)


                print("\n")

            except Exception as e:
                print(f"Error processing one email: {e}\n")
                traceback.print_exc()
                skipped += 1

        # Save if needed
        if wrote > 0 and not DRY_RUN:
            wb.Save()

        print(f"\nSummary: wrote {wrote}, skipped {skipped}")

    finally:
        try:
            if ALWAYS_LEAVE_WORKBOOK_OPEN:
                pass
            else:
                if wb is not None and opened_here:
                    wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if ALWAYS_LEAVE_EXCEL_RUNNING:
                pass
            else:
                if excel is not None and not excel_was_running:
                    excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


if __name__ == "__main__":
    try:
        process_claims()
    except Exception as e:
        print("Fatal error:", e)
        traceback.print_exc()
        sys.exit(1)
