# SlipSender.py
# Creates one Outlook draft per District with all matching packing slips attached
# and an HTML table of F–N cells from "Recieved Materials Template".
# Row selection: F not blank, starts with 904 or TD, and F cell has no fill.
# After drafting, fills Column F yellow. Logs each row to an .xlsx log.

# Migrated to config-driven architecture.
# Now reads all settings from config/config.py instead of hardcoded paths.

import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

# Import config from parent directories
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "config"))
import config

# ---- Settings from config (now portable) ----
WORKBOOK_PATH = config.SLIPSENDER_SETTINGS["workbook_path"]
SHEET_NAME = config.SLIPSENDER_SETTINGS["worksheet_name"]
DISTRICT_MAP_PATH = config.SLIPSENDER_SETTINGS["district_map_file"]
PACKING_SLIPS_ROOT = config.SLIPSENDER_SETTINGS["packing_slips_root"]
ALLOWED_EXTS = config.SLIPSENDER_SETTINGS["allowed_extensions"]
ALWAYS_CC = config.SLIPSENDER_SETTINGS["always_cc"]
LOG_FOLDER = config.SLIPSENDER_SETTINGS["log_folder"]
LOG_FILENAME = config.SLIPSENDER_SETTINGS["log_filename"]
YELLOW_HEX = config.SLIPSENDER_SETTINGS["yellow_hex"]
MISSING_ROW_BG = config.SLIPSENDER_SETTINGS["missing_row_bg"]
LOG_XLSX = os.path.join(LOG_FOLDER, LOG_FILENAME)

# Greeting / body top
BODY_GREETING = """Hello,<br>
Wesco has delivered materials for the following work orders.<br>
I have attached the packing slips for your records.<br><br><br>
Thank you.<br><br>
"""

# ---- Utility Functions ----

def ensure_dirs(path):
    d = os.path.dirname(path)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def rgb_from_hex(hex_str):
    h = hex_str.lstrip("#")
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    return b + (g << 8) + (r << 16)


def find_packing_slips(wo, root, allowed_exts):
    wo_l = str(wo).lower()
    matches = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            ext = os.path.splitext(fn)[1].lower()
            if ext in allowed_exts and wo_l in fn.lower():
                matches.append(os.path.join(dirpath, fn))
    return matches


def clean_semicolons(s):
    if not s:
        return ""
    s = s.replace(",", ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]
    return "; ".join(parts)


def build_table_html(rows, missing_set):
    cols = [
        ("F", "Work Order #"),
        ("G", "District"),
        ("H", "WESCO PO"),
        ("I", "Fixture SAP"),
        ("J", "Qty"),
        ("K", "Fixture Description"),
        ("L", "Photocell SAP"),
        ("M", "QTY"),
        ("N", "Description"),
    ]
    th_style = "padding:6px 8px;border:1px solid #444;font-weight:600;text-align:left;"
    td_style = "padding:6px 8px;border:1px solid #777;text-align:left;"

    html = []
    html.append('<div style="font-family:Segoe UI, Arial, sans-serif;font-size:12px;">')
    html.append('<table cellpadding="0" cellspacing="0" style="border-collapse:collapse;width:100%;">')
    html.append("<tr>")
    for _, header in cols:
        html.append(f'<th style="{th_style}">{header}</th>')
    html.append("</tr>")

    for r in rows:
        wo = str(r.get("F", "")).strip()
        row_bg = f'background-color:{MISSING_ROW_BG};' if wo in missing_set else ""
        html.append(f'<tr style="{row_bg}">')
        for key, _ in cols:
            val = r.get(key, "")
            html.append(f'<td style="{td_style}">{val}</td>')
        html.append("</tr>")

    html.append("</table></div>")
    return "".join(html)


def valid_wo_cell_text(txt):
    if not txt:
        return False
    s = str(txt).strip()
    return s.upper().startswith("TD") or s.startswith("904")


def normalize_id_value(v):
    """Normalize ID-like numbers (WO, PO, SAPs)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main():
    ensure_dirs(LOG_XLSX)

    import win32com.client as win32
    from openpyxl import Workbook, load_workbook

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = True   # keep Excel visible
    excel.DisplayAlerts = False

    wb = None
    try:
        print("📂 Opening workbook...")
        wb = excel.Workbooks.Open(WORKBOOK_PATH)
        ws = wb.Worksheets(SHEET_NAME)

        used = ws.UsedRange
        first_row = used.Row
        n_rows = used.Rows.Count
        last_row = first_row + n_rows - 1

        col_map = {"F": 6, "G": 7, "H": 8, "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14}
        xlColorIndexNone = -4142

        candidates = []
        for r in range(2, last_row + 1):
            f_cell = ws.Cells(r, col_map["F"])
            f_val = f_cell.Value
            if f_val is None or str(f_val).strip() == "":
                continue

            f_val_str = normalize_id_value(f_val)
            if not valid_wo_cell_text(f_val_str):
                continue

            try:
                no_fill = (f_cell.Interior.ColorIndex == xlColorIndexNone)
            except Exception:
                no_fill = True
            if not no_fill:
                continue

            row_data = {}
            for k, c in col_map.items():
                v = ws.Cells(r, c).Value
                if k in ("F", "H", "I", "L"):
                    row_data[k] = normalize_id_value(v)
                else:
                    row_data[k] = "" if v is None else str(v).strip()
            row_data["_row_index"] = r
            candidates.append(row_data)

        if not candidates:
            print("⚠️ No eligible rows found.")
            return

        print(f"✅ Found {len(candidates)} eligible rows.")

        district_wb = load_workbook(DISTRICT_MAP_PATH, data_only=True)
        dws = district_wb.worksheets[0]  # First sheet
        district_map = {}

        for row in dws.iter_rows(min_row=2, values_only=True):
            dist = (row[0] or "")
            if not dist:
                continue
            emails = []
            for idx in range(1, min(len(row), 6)):  # B–F
                val = row[idx]
                if val:
                    emails.append(str(val).strip())
            to_all = "; ".join(emails)
            district_map[dist.upper()] = {
                "to": clean_semicolons(to_all),
                "cc": ""  # only ALWAYS_CC appended later
            }

        groups = {}
        for r in candidates:
            dist = (r.get("G") or "").strip()
            groups.setdefault(dist.upper(), []).append(r)

        outlook = win32.Dispatch("Outlook.Application")

        if os.path.exists(LOG_XLSX):
            log_wb = load_workbook(LOG_XLSX)
            log_ws = log_wb.active
        else:
            log_wb = Workbook()
            log_ws = log_wb.active
            log_ws.title = "Log"
            log_ws.append(["Timestamp", "District", "WO#", "Attachments", "Missing (Y/N)", "Draft EntryID", "Subject"])

        yellow_rgb = rgb_from_hex(YELLOW_HEX)
        total_drafts = 0

        for dist_key, rows in groups.items():
            print(f"\n📍 Processing district: {dist_key} ({len(rows)} rows)")
            rec = district_map.get(dist_key)
            if not rec:
                print(f"⚠️ District '{dist_key}' not found in District.xlsx; skipping.")
                continue
            to_addr = rec["to"]
            cc_addr = rec["cc"]
            if ALWAYS_CC:
                cc_addr = "; ".join([p for p in [cc_addr, ALWAYS_CC] if p]).strip("; ").strip()

            attachments = []
            missing_wos = set()
            wo_list = []

            for r in rows:
                wo = (r.get("F") or "").strip()
                if not wo:
                    continue
                print(f"   ➡️ Work Order {wo} ...", end="")
                found = find_packing_slips(wo, PACKING_SLIPS_ROOT, ALLOWED_EXTS)
                if found:
                    print(f" found {len(found)} file(s).")
                    attachments.extend(found)
                else:
                    print(" no files found!")
                    missing_wos.add(wo)
                wo_list.append(wo)

            table_html = build_table_html(rows, missing_wos)
            mail = outlook.CreateItem(0)
            mail.Display(False)
            sig_html = mail.HTMLBody or ""
            subject = "Packing Slips - " + ", ".join(wo_list)
            body_html = BODY_GREETING + table_html + "<br><br>" + sig_html
            mail.HTMLBody = body_html
            mail.Subject = subject
            mail.To = to_addr
            mail.CC = cc_addr

            for fpath in attachments:
                try:
                    mail.Attachments.Add(fpath)
                except Exception as e:
                    print(f"     ⚠️ Failed to attach {fpath}: {e}")

            mail.Save()
            entry_id = mail.EntryID
            print(f"✅ Draft created for district {dist_key} with {len(wo_list)} WO(s), "
                  f"{len(attachments)} attachment(s), {len(missing_wos)} missing.")

            for r in rows:
                row_idx = r["_row_index"]
                f_cell = ws.Cells(row_idx, col_map["F"])
                try:
                    f_cell.Interior.Color = yellow_rgb
                except Exception:
                    pass

            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for r in rows:
                wo = (r.get("F") or "").strip()
                these_files = find_packing_slips(wo, PACKING_SLIPS_ROOT, ALLOWED_EXTS)
                log_ws.append([ts, r.get("G", ""), wo, "; ".join(these_files),
                               "Y" if wo in missing_wos else "N", entry_id, subject])

            total_drafts += 1

        wb.Save()
        log_wb.save(LOG_XLSX)
        print(f"\n✅ Done. Drafts created: {total_drafts}. Log saved to {LOG_XLSX}")

    except Exception:
        traceback.print_exc()
    finally:
        # Leave Excel and workbook open
        pass


if __name__ == "__main__":
    main()
