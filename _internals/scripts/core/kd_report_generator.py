"""
Dual Report Generator
=====================
Generates two reports from tracking workbook:
1. New Development Report (TD worksheet, filter AF blank)
2. Car Hit Pole Report (MainSheet worksheet, filter AE blank)

Emails both reports to all whitelist recipients.
Uses configuration from _internals/config/config.py
"""

import pandas as pd
import win32com.client as win32
from datetime import datetime
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment, Font, PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule

# Import config from parent directories
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "config"))
import config

# Column mappings for New Development Report (TD worksheet)
NEW_DEV_COLUMN_MAPPING = {
    'C': 'District',
    'D': 'Work Order #',
    'I': 'WO Received Date',
    'K': 'Service Pricing Submitted',
    'Y': 'Civil Release Date',
    'R': 'PO Received Date',
    'AA': 'Fixture ETA from Wesco',
    'AB': 'Pole ETA From Ameron',
    'AE': 'Foundation Installation Date',
    'J': 'Pole Installation Date',
    'W': 'Status',
    'CJ': 'Age',
    'U': 'Permit Approved (Y/N)',
    'V': 'Permit Expiration',
    'T': 'TD POLE QTY'
}

# Column mappings for Car Hit Pole Report (MainSheet worksheet)
CAR_HIT_POLE_COLUMN_MAPPING = {
    'C': 'District',
    'D': 'Work Order #',
    'I': 'Sasco WO Received Date',
    'K': 'Service Pricing Submitted',
    'R': 'PO Received Date',
    'Z': 'Fixture ETA From Wesco',
    'AA': 'Pole ETA From Ameron',
    'J': 'Pole Installation Date',
    'W': 'Status',
    'CI': 'Age',
    'U': 'Permit Approved (y/n)',
    'V': 'Permit Expiration',
    'BW': 'Decorative Fixture',
    'BX': 'Department Needing Action'
}

# Configuration from config.py
WORKBOOK_PATH = config.WOP22_SETTINGS["excel_path"]
OUTPUT_FOLDER = os.path.join(config.DATA_PATHS["output"], "kd_report_output")
EMAIL_RECIPIENTS = config.OUTLOOK_SETTINGS["whitelist_emails"]  # All whitelist recipients
MAILBOX_NAME = config.OUTLOOK_SETTINGS["mailbox_name"]

def col_letter_to_index(letter):
    """Convert Excel column letter to 0-based index"""
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def read_and_filter_data(worksheet_name, filter_column):
    """Read worksheet using Excel COM (works with open files) and filter for rows where specified column is blank"""
    print(f"Reading workbook: {WORKBOOK_PATH}")
    print(f"Worksheet: {worksheet_name}")
    print("Using Excel COM automation (file can be open)...")

    excel = None
    workbook = None
    workbook_was_open = False

    try:
        # Connect to Excel application (use existing instance if available)
        try:
            excel = win32.GetActiveObject('Excel.Application')
        except:
            excel = win32.Dispatch('Excel.Application')

        excel.Visible = True  # Keep Excel visible so workbook stays open
        excel.DisplayAlerts = False

        # Check if workbook is already open
        for wb in excel.Workbooks:
            if wb.FullName.lower() == WORKBOOK_PATH.lower():
                workbook = wb
                workbook_was_open = True
                print("Found workbook already open in Excel")
                break

        # If not open, open it
        if workbook is None:
            print("Opening workbook...")
            workbook = excel.Workbooks.Open(WORKBOOK_PATH, ReadOnly=False)

        # Access the worksheet
        worksheet = workbook.Sheets(worksheet_name)

        # Get the used range
        used_range = worksheet.UsedRange
        data = used_range.Value

        # Convert to pandas DataFrame
        # First row is headers
        if data and len(data) > 1:
            headers = data[0]
            rows = data[1:]
            # Create DataFrame with dtype=object to prevent pandas from inferring datetime64 with timezone
            df = pd.DataFrame(rows, columns=headers, dtype=object)
        else:
            df = pd.DataFrame()

        # Clean up datetime values: strip any timezone info from Python datetime objects
        # This must be done before pandas converts them to datetime64
        for col in df.columns:
            # Check first non-null value to see if it's a datetime
            first_val = df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else None
            if isinstance(first_val, (datetime, pd.Timestamp)):
                # Strip timezone from all datetime values in this column
                df[col] = df[col].apply(lambda x:
                    x.replace(tzinfo=None) if hasattr(x, 'tzinfo') and x.tzinfo is not None else x
                )

        print(f"Total rows in source: {len(df)}")

        # Filter: Keep only rows where specified column is blank
        filter_index = col_letter_to_index(filter_column)
        df_filtered = df[df.iloc[:, filter_index].isna()]

        print(f"Rows after filtering ({filter_column} blank): {len(df_filtered)}")

        return df_filtered

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise

    finally:
        # Don't close the workbook - leave it open for user
        # Don't quit Excel
        pass

def transform_data(df_filtered, column_mapping):
    """Select and rename columns according to mapping"""
    print("Transforming data...")

    # Create a new dataframe with selected columns and new headers
    new_data = {}
    for source_col, dest_header in column_mapping.items():
        col_index = col_letter_to_index(source_col)
        new_data[dest_header] = df_filtered.iloc[:, col_index].values

    # Create DataFrame with dtype=object to prevent datetime64 timezone inference
    df_report = pd.DataFrame(new_data, dtype=object)

    print(f"Output columns: {len(df_report.columns)}")
    print(f"Output rows: {len(df_report)}")

    return df_report

def save_new_development_report(df_report):
    """Save New Development Report to output folder with dated filename and specific formatting"""
    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Generate filename with current date (using hyphens instead of slashes for valid filename)
    today = datetime.now().strftime('%m-%d-%y')
    filename = f"New Development Report {today}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, filename)

    # Remove timezone information by processing each cell individually
    # This is the most reliable way to handle timezone-aware datetime objects from Excel COM
    import copy
    df_clean = df_report.copy()

    for col in df_clean.columns:
        df_clean[col] = df_clean[col].apply(lambda x:
            x.replace(tzinfo=None) if isinstance(x, pd.Timestamp) and x.tzinfo is not None
            else (x.replace(tzinfo=None) if hasattr(x, 'tzinfo') and x.tzinfo is not None else x)
        )

    print(f"Saving report to: {output_path}")
    df_clean.to_excel(output_path, index=False)

    # Apply formatting to the Excel file
    print("Applying formatting...")
    wb = load_workbook(output_path)
    ws = wb.active

    # Columns to format as Short Date: C, D, E, F, G, H, I, J, M, N
    date_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'M', 'N']

    # Define styles
    center_alignment = Alignment(horizontal='center', vertical='center')
    center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')  # Grey-blue color

    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Center align all cells (with wrap text for Status column K)
            if cell.column_letter == 'K':
                cell.alignment = center_wrap_alignment
            else:
                cell.alignment = center_alignment

            # Header row formatting (row 1)
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill

            # Apply Short Date format to specific columns (skip header)
            if cell.column_letter in date_columns and cell.row > 1:
                cell.number_format = 'm/d/yyyy'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set column width with some padding and 20% extra for filter buttons
        adjusted_width = (max_length + 2) * 1.20

        # Make District column (A) an additional 15% wider
        if column_letter == 'A':
            adjusted_width = adjusted_width * 1.15

        # Make Age column (L) an additional 40% wider
        if column_letter == 'L':
            adjusted_width = adjusted_width * 1.40

        # Make Status column (K) 1/3 of its calculated width
        if column_letter == 'K':
            adjusted_width = adjusted_width / 3

        ws.column_dimensions[column_letter].width = adjusted_width

    # Set row height to 28 for all rows
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 28

    # Freeze panes at C2 (freezes columns A-B horizontally and header row vertically)
    ws.freeze_panes = 'C2'

    # Apply AutoFilter to all columns
    ws.auto_filter.ref = ws.dimensions

    # Add conditional formatting to Age column (L) - Color scale: Red (high) -> Yellow (mid) -> Green (low)
    age_range = f'L2:L{ws.max_row}'
    color_scale = ColorScaleRule(
        start_type='min',
        start_color='63BE7B',  # Green
        mid_type='percentile',
        mid_value=50,
        mid_color='FFEB84',  # Yellow
        end_type='max',
        end_color='F8696B'  # Red
    )
    ws.conditional_formatting.add(age_range, color_scale)

    # Save the formatted workbook
    wb.save(output_path)
    wb.close()

    print("New Development Report saved successfully with formatting!")

    return output_path, filename

def save_car_hit_pole_report(df_report):
    """Save Car Hit Pole Report to output folder with dated filename and specific formatting"""
    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Generate filename with current date
    today = datetime.now().strftime('%m-%d-%y')
    filename = f"Car Hit Pole Report {today}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, filename)

    # Aggressive timezone stripping - iterate through ALL cells
    # Since DataFrame is dtype=object, we can safely iterate and modify
    df_clean = df_report.copy()

    # Strip timezone from every cell that is a datetime object
    for col in df_clean.columns:
        df_clean[col] = df_clean[col].apply(lambda x:
            x.replace(tzinfo=None) if hasattr(x, 'tzinfo') and x.tzinfo is not None else x
        )

    print(f"Saving report to: {output_path}")
    df_clean.to_excel(output_path, index=False)

    # Apply formatting
    print("Applying formatting...")
    wb = load_workbook(output_path)
    ws = wb.active

    # Date columns for Car Hit Pole: C, D, E, F, G, H, L (output positions)
    date_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'L']

    # Define styles
    center_alignment = Alignment(horizontal='center', vertical='center')
    center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Center align all cells (with wrap text for Status and Permit Approved columns)
            if cell.column_letter in ['I', 'K']:  # Status (I) and Permit Approved (K) columns
                cell.alignment = center_wrap_alignment
            else:
                cell.alignment = center_alignment

            # Header row formatting
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill

            # Apply Short Date format to specific columns
            if cell.column_letter in date_columns and cell.row > 1:
                cell.number_format = 'm/d/yyyy'

            # Apply General format to Age column (J) to fix inconsistent formatting
            if cell.column_letter == 'J' and cell.row > 1:
                cell.number_format = 'General'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set column width with 20% extra for filter buttons
        adjusted_width = (max_length + 2) * 1.20

        # District column (A) - additional 15% wider
        if column_letter == 'A':
            adjusted_width = adjusted_width * 1.15

        # Age column (J) - reduce width by 30% (70% of base width)
        if column_letter == 'J':
            adjusted_width = adjusted_width * 0.70

        # Status column (I) - 1/3 width
        if column_letter == 'I':
            adjusted_width = adjusted_width / 3

        # Permit Approved column (K) - same width as Status column (I)
        if column_letter == 'K':
            adjusted_width = adjusted_width / 3

        ws.column_dimensions[column_letter].width = adjusted_width

    # Set row height to 28
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 28

    # Freeze panes at C2
    ws.freeze_panes = 'C2'

    # Apply AutoFilter
    ws.auto_filter.ref = ws.dimensions

    # Conditional formatting for Age column (J) - Color scale
    age_range = f'J2:J{ws.max_row}'
    color_scale = ColorScaleRule(
        start_type='min',
        start_color='63BE7B',  # Green
        mid_type='percentile',
        mid_value=50,
        mid_color='FFEB84',  # Yellow
        end_type='max',
        end_color='F8696B'  # Red
    )
    ws.conditional_formatting.add(age_range, color_scale)

    # Save the formatted workbook
    wb.save(output_path)
    wb.close()

    print("Car Hit Pole Report saved successfully with formatting!")

    return output_path, filename

def send_email(attachment_paths, subject):
    """Create draft email in Outlook with multiple attachments (not sent automatically)"""
    print("\nCreating draft email in Outlook...")

    # Join all recipients with semicolons
    recipients = "; ".join(EMAIL_RECIPIENTS)
    print(f"To: {recipients}")
    print(f"Subject: {subject}")

    try:
        # Create Outlook application object
        outlook = win32.Dispatch('Outlook.Application')

        # Create a new mail item
        mail = outlook.CreateItem(0)  # 0 = MailItem

        # Set email properties
        mail.To = recipients
        mail.Subject = subject
        mail.Body = subject  # Body same as subject

        # Attach all reports
        for attachment_path in attachment_paths:
            mail.Attachments.Add(attachment_path)
            print(f"Attached: {os.path.basename(attachment_path)}")

        # Display the draft (does not send)
        mail.Display()

        print(f"Draft email created successfully!")
        print("Review the email and click Send when ready.")

    except Exception as e:
        print(f"Error creating draft email: {e}")
        raise

def main():
    """Main execution function - generates both reports and emails them"""
    try:
        print("=" * 70)
        print("Dual Report Generator")
        print("=" * 70)
        print()

        attachment_paths = []

        # ===== REPORT 1: NEW DEVELOPMENT REPORT =====
        print("\n" + "=" * 70)
        print("Generating Report 1: New Development Report (TD Worksheet)")
        print("=" * 70)

        # Read and filter TD worksheet (column AF blank)
        df_filtered_td = read_and_filter_data("TD", "AF")

        # Transform data using New Development column mapping
        df_new_dev = transform_data(df_filtered_td, NEW_DEV_COLUMN_MAPPING)

        # Save New Development Report
        new_dev_path, new_dev_filename = save_new_development_report(df_new_dev)
        attachment_paths.append(new_dev_path)

        # ===== REPORT 2: CAR HIT POLE REPORT =====
        print("\n" + "=" * 70)
        print("Generating Report 2: Car Hit Pole Report (MainSheet Worksheet)")
        print("=" * 70)

        # Read and filter MainSheet worksheet (column AE blank)
        df_filtered_main = read_and_filter_data("MainSheet", "AE")

        # Transform data using Car Hit Pole column mapping
        df_car_hit = transform_data(df_filtered_main, CAR_HIT_POLE_COLUMN_MAPPING)

        # Save Car Hit Pole Report
        car_hit_path, car_hit_filename = save_car_hit_pole_report(df_car_hit)
        attachment_paths.append(car_hit_path)

        # ===== EMAIL BOTH REPORTS =====
        print("\n" + "=" * 70)
        print("Emailing Both Reports")
        print("=" * 70)

        # Create email subject with today's date
        today = datetime.now().strftime('%m-%d-%y')
        subject = f"Reports for {today}"

        # Send email with both attachments
        send_email(attachment_paths, subject)

        print()
        print("=" * 70)
        print("BOTH REPORTS GENERATED SUCCESSFULLY!")
        print("=" * 70)
        print(f"Report 1: {new_dev_filename}")
        print(f"Report 2: {car_hit_filename}")
        print("=" * 70)

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to exit...")
        sys.exit(1)

if __name__ == "__main__":
    main()
