#!/usr/bin/env python3
r"""
JIS Page Populator — v2.0

Migrated to config-driven architecture.
Now reads all settings from config/config.py instead of hardcoded paths.

Delta from v1.3.1:
- Two Excel logs (single-sheet 'Log'): Preflight first, then Final after outputs.
- De-dup: if no standard LED wattage exists, ONLY log "No standard wattage defined"
  (do NOT also log the removal-unknown line).
- Log 'Type' labels updated:
    Mismatch
    Blank install wattage, default to conversion standard
    No standard wattage defined
    Unable to build description in output file
    SKIPPED OUTPUT - NO INSTALL DATE
- LED detection for removal path: ONLY when Input col B **ends with 'W'** (spaces allowed before W).
- Normalize conversion keys by UPPER, removing spaces and commas (e.g., '22,000 L' → '22000L').
- Lamp Size in logs is the raw input text (blank if actually empty), not 'nan'.

Unchanged (recap):
- Input (first sheet, positional): A=Structure, B=Lamp Size, C=LED Change Date, D=LED Wattage, E=Comments(optional)
- Conversion (first sheet, by letters): A=Lamp key, B=HPSV, C=LED standard, D=MH(optional)
- Setup (first sheet): A2→Y1 (with " - VARIOUS LOCATIONS"), B2→AF5, C2→AO35 (mm/dd/yyyy), D2:D→exclusions
- Output workbook per date "( JIS ) JOB INSTRUCTION SHEET <MM.DD.YYYY>.xlsx"
- Page capacity: 11 structures/page (rows 11–32 as R/I pairs)
- Highlight (M col on both R & I) when Installed(D) present and ≠ Standard(C), unless excluded
"""
import sys, math, subprocess, datetime as dt, re, argparse
from pathlib import Path

# Import config from parent directories
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "config"))
import config

import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---- Paths from config (now portable) ----
INPUT_DIR   = Path(config.JIS_SETTINGS["input_folder"])
OUTPUT_DIR  = Path(config.JIS_SETTINGS["output_folder"])
LOG_DIR     = Path(config.JIS_SETTINGS["log_folder"])
SETUP_XLSX  = Path(config.JIS_SETTINGS["setup_file"])
CONV_DIR    = Path(config.JIS_SETTINGS["conversion_chart_folder"])
CONV_XLSX   = Path(config.JIS_SETTINGS["conversion_chart_file"])
TEMPLATE_X  = Path(config.JIS_SETTINGS["template_file"])

# ---- Template geometry ----
LINE_START_ROW = 11
LINE_END_ROW   = 32   # inclusive
LINES_PER_PAGE = (LINE_END_ROW - LINE_START_ROW + 1)  # 22
STRUCTS_PER_PAGE = LINES_PER_PAGE // 2  # 11 (R then I)

# Columns / cells
COL_STRUCTURE = "B"
COL_LEGEND    = "J"
COL_DESC      = "M"
COL_QTY_I     = "AJ"
COL_QTY_R     = "AP"

CELL_DATE     = "T3"
CELL_HEADER   = "Y1"
CELL_PAGE_CUR = "AR1"
CELL_PAGE_TOT = "AU1"
CELL_AF5      = "AF5"
CELL_AO35     = "AO35"

__version__ = "2.0"

def ensure_deps():
    for p in ("pandas","openpyxl","pywin32"):
        try:
            __import__("win32com.client" if p=="pywin32" else p)
        except ImportError:
            print(f"[info] Installing missing dependency: {p} ...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", p])

ensure_deps()

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage

# LED only when Lamp Size ENDS WITH W (spaces allowed before 'W')
ENDSWITH_W_RE = re.compile(r"^\s*.*?(\d+(?:\.\d+)?)\s*W\s*$", re.IGNORECASE)

def normalize_key(raw: str) -> str:
    """Uppercase; remove spaces and commas."""
    if raw is None:
        return ""
    s = str(raw).upper().replace(" ", "").replace(",", "").strip()
    return s

def clean_watts(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip().upper().replace("W","")
    return "".join(ch for ch in s if ch.isdigit() or ch==".")

def safe_display(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s

def find_first_xlsx(folder: Path):
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {folder}")
    return files[0]

def read_first_sheet_as_df(path: Path):
    xl = pd.ExcelFile(path)
    first_sheet = xl.sheet_names[0]
    return pd.read_excel(path, sheet_name=first_sheet, dtype=str)

def parse_input_positional(df: pd.DataFrame, assume_positional: bool):
    if df.shape[1] < 4:
        raise ValueError("Input sheet must have at least 4 columns (A-D).")
    # Friendly warning if headers look off (can bypass with --assume-positional)
    cols5 = list(df.columns[:5])
    found_headers = [str(c) for c in cols5]
    expected_text = ["Structure Number","Lamp Size","LED Change Date","LED Wattage","Comments (optional)"]
    norm = [str(c).strip().lower() for c in found_headers]
    expected_norm = ["structure", "lamp", "date", "led", "comments"]
    ok = True
    for i, n in enumerate(norm[:4]):
        if expected_norm[i] not in n:
            ok = False
    if not ok and not assume_positional:
        print("[JIS] WARNING: Input headers differ from typical names.")
        print("  Expected (by meaning):", ", ".join(expected_text))
        print("  Found (first 5 cols) :", ", ".join(found_headers))
        yn = input("Proceed using positional columns A..E anyway? (Y/N): ").strip().lower()
        if yn not in ("y","yes"):
            sys.exit("[JIS] Aborted by user due to header mismatch.")

    new_cols = ["Structure Number","Lamp Size","LED Change Date","LED Wattage"]
    data = df.iloc[:, :4].copy()
    data.columns = new_cols
    comments_series = ""
    if df.shape[1] >= 5:
        comments_series = df.iloc[:,4]
    data["Comments"] = comments_series if isinstance(comments_series, pd.Series) else ""
    return data

def parse_input_and_group_by_date(df_in: pd.DataFrame, assume_positional: bool):
    df = parse_input_positional(df_in, assume_positional)
    out = df.copy()
    out["_parsed_date"] = pd.to_datetime(out["LED Change Date"], errors="coerce")
    no_date_mask = out["_parsed_date"].isna()
    with_date = out[~no_date_mask].copy()
    skipped_no_date = out[no_date_mask].copy()

    with_date["_LED_W"]  = with_date["LED Wattage"].apply(clean_watts)
    with_date["_Struct"] = with_date["Structure Number"].astype(str).str.strip()
    # Keep raw for logging/UI; also build normalized key for lookups
    with_date["_LampRaw"] = with_date["Lamp Size"]  # keep original as-is
    with_date["_LampKey"] = with_date["Lamp Size"].apply(normalize_key)
    with_date["_Comments"] = with_date["Comments"].fillna("")

    with_date["_date_display"] = with_date["_parsed_date"].dt.strftime("%m/%d/%Y")
    with_date["_date_key"]     = with_date["_parsed_date"].dt.strftime("%m.%d.%Y")

    groups = {dk: g.reset_index(drop=True) for dk,g in with_date.groupby("_date_key", sort=True)}
    return groups, skipped_no_date[["Structure Number","Lamp Size","LED Change Date","LED Wattage","Comments"]]

def load_setup_fields_and_exclusions():
    wb = load_workbook(SETUP_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    a2 = ws["A2"].value
    b2 = ws["B2"].value
    c2 = ws["C2"].value
    excl = set()
    for r in range(2, ws.max_row + 1):
        val = ws[f"D{r}"].value
        if val is None:
            continue
        s = str(val).strip()
        if s:
            excl.add(s)
    a2 = "" if a2 is None else str(a2).strip()
    b2 = "" if b2 is None else str(b2).strip()
    return a2, b2, c2, excl

def load_conversion_maps():
    xl = pd.ExcelFile(CONV_XLSX)
    sheet = xl.sheet_names[0]
    conv = pd.read_excel(CONV_XLSX, sheet_name=sheet, dtype=str, header=0)
    ncols = conv.shape[1]
    if ncols < 5:
        raise ValueError("Conversion chart must have at least columns A, B, C, D, E (Lamp, HPSV, LED, MH, LPSV).")
    lamp_series = conv.iloc[:, 0]  # A
    hpsv_series = conv.iloc[:, 1]  # B
    led_series  = conv.iloc[:, 2]  # C
    mh_series   = conv.iloc[:, 3] if ncols >= 4 else None  # D (optional)
    lpsv_series = conv.iloc[:, 4] if ncols >= 5 else None  # E (new LPSV column)

    def norm_key(x):
        return normalize_key(x)

    hpsv_map, led_std_map, mh_map, lpsv_map = {}, {}, {}, {}
    for i in range(len(conv)):
        lk = norm_key(lamp_series.iat[i] if i < len(lamp_series) else "")
        if not lk:
            continue
        hv = clean_watts(hpsv_series.iat[i] if i < len(hpsv_series) else "")
        lv = clean_watts(led_series.iat[i] if i < len(led_series) else "")
        mv = clean_watts(mh_series.iat[i]) if mh_series is not None and i < len(mh_series) else ""

        if hv and lk not in hpsv_map:
            hpsv_map[lk] = hv
        if lv and lk not in led_std_map:
            led_std_map[lk] = lv
        if mv and lk not in mh_map:
            mh_map[lk] = mv

        lpsv = clean_watts(lpsv_series.iat[i]) if lpsv_series is not None and i < len(lpsv_series) else ""
        if lpsv and lk not in lpsv_map:
            lpsv_map[lk] = lpsv

    return hpsv_map, mh_map, led_std_map, lpsv_map

def timestamp_for_filename():
    return dt.datetime.now().strftime("%m.%d.%Y.%H.%M")

def ensure_dirs():
    for p in (OUTPUT_DIR, LOG_DIR):
        p.mkdir(parents=True, exist_ok=True)

def log_path(kind: str):
    return LOG_DIR / f"JIS_Log_{kind}_{timestamp_for_filename()}.xlsx"

# ---------- Build a log (no output writing) ----------
def build_log_rows(groups, skipped_no_date_df, exclusions_set, hpsv_map, mh_map, led_std_map, lpsv_map):
    rows = []
    lvps = lpsv_map  # alias to guard against stray 'lvps' typos

    # SKIPPED (no date)
    if not skipped_no_date_df.empty:
        for _, r in skipped_no_date_df.iterrows():
            struct = safe_display(r.get("Structure Number",""))
            if struct in exclusions_set:
                continue
            lamp_raw = safe_display(r.get("Lamp Size",""))
            cmt      = safe_display(r.get("Comments","")) or "No comments"
            rows.append({
                "Type":"SKIPPED OUTPUT - NO INSTALL DATE",
                "Structure Number":struct,
                "Lamp Size":lamp_raw,
                "Date":"",
                "Standard LED":"",
                "Installed LED":"",
                "Comments":cmt
            })

    # Per-date discrepancies
    for date_key, g in groups.items():
        date_display = g["_date_display"].iloc[0]
        for _, r in g.iterrows():
            struct_no = safe_display(r["_Struct"])
            if struct_no in exclusions_set:
                continue
            lamp_raw  = safe_display(r["_LampRaw"])
            lamp_key  = r["_LampKey"] or ""
            installed = clean_watts(r["_LED_W"])
            cmt       = safe_display(r["_Comments"]) or "No comments"

            # Determine expected wattage for standards
            exp_led = led_std_map.get(lamp_key, "")
            exp_lpsv = lpsv_map.get(lamp_key, "")

            # Discrepancy logging uses the LED standard (not the LPSV removal equivalent)
            if exp_led:
                if installed == "":
                    rows.append({
                        "Type": "Blank install wattage, default to conversion standard",
                        "Structure Number": struct_no,
                        "Lamp Size": lamp_raw,
                        "Date": date_display,
                        "Standard LED": exp_led,
                        "Installed LED": "",
                        "Comments": cmt
                    })
                elif installed != exp_led:
                    rows.append({
                        "Type": "Mismatch",
                        "Structure Number": struct_no,
                        "Lamp Size": lamp_raw,
                        "Date": date_display,
                        "Standard LED": exp_led,
                        "Installed LED": installed,
                        "Comments": cmt
                    })
            else:
                # No LED standard defined for this key
                rows.append({
                    "Type": "No standard wattage defined",
                    "Structure Number": struct_no,
                    "Lamp Size": lamp_raw,
                    "Date": date_display,
                    "Standard LED": "",
                    "Installed LED": installed,
                    "Comments": cmt
                })

    return rows

def write_excel_log(rows, kind: str):
    if not rows:
        print(f"[JIS] No discrepancies to log for {kind}.")
        return None
    ensure_dirs()
    out = log_path(kind)
    cols = ["Type","Structure Number","Lamp Size","Date","Standard LED","Installed LED","Comments"]
    df = pd.DataFrame(rows, columns=cols)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="Log", index=False)
    print(f"[JIS v{__version__}] Wrote {kind} log: {out}")
    return out

# ----------------- COM OUTPUTS -----------------
def generate_outputs_com(groups, header_a2, af5_b2, ao35_c2, exclusions_set,
                         hpsv_map, mh_map, led_std_map, lpsv_map):
    try:
        import win32com.client as win32, pywintypes
        lvps = lpsv_map  # alias to guard against stray 'lvps' typos
        import win32com.client as win32, pywintypes
    except Exception as e:
        print("[JIS] COM mode not available:", e)
        return False

    excel = None
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        ensure_dirs()

        for date_key, g in groups.items():
            date_display = g["_date_display"].iloc[0]
            structs = g[["_Struct","_LampKey","_LampRaw","_LED_W","_Comments"]].values.tolist()
            pages = math.ceil(len(structs) / STRUCTS_PER_PAGE)

            wb = excel.Workbooks.Open(str(TEMPLATE_X))
            base_ws = wb.Sheets(1)
            base_ws.Name = "Sheet1"
            while wb.Sheets.Count < pages:
                base_ws.Copy(After=wb.Sheets(wb.Sheets.Count))
            for i in range(1, pages+1):
                wb.Sheets(i).Name = f"Sheet{i}"
            while wb.Sheets.Count > pages:
                wb.Sheets(wb.Sheets.Count).Delete()

            # AO35 date
            ao35_dt = None
            try:
                dt_val = pd.to_datetime(ao35_c2, errors="coerce")
                if pd.notna(dt_val):
                    ao35_dt = dt_val.to_pydatetime()
            except Exception:
                ao35_dt = None
            ao35_com = None
            if ao35_dt is not None:
                import pywintypes as pwt
                ao35_com = pwt.Time(ao35_dt)

            for p_idx in range(1, pages+1):
                ws = wb.Sheets(p_idx)
                ws.Range(CELL_DATE).Value = date_display
                ws.Range(CELL_HEADER).Value = f"{header_a2} - VARIOUS LOCATIONS"
                ws.Range(CELL_PAGE_CUR).Value = p_idx
                ws.Range(CELL_PAGE_TOT).Value = pages
                ws.Range(CELL_AF5).Value = af5_b2
                if ao35_com is not None:
                    ws.Range(CELL_AO35).Value = ao35_com
                    ws.Range(CELL_AO35).NumberFormat = "mm/dd/yyyy"
                else:
                    ws.Range(CELL_AO35).Value = ""

                # clear grid
                for col in (COL_STRUCTURE, COL_LEGEND, COL_QTY_I, COL_QTY_R, COL_DESC):
                    ws.Range(f"{col}{LINE_START_ROW}:{col}{LINE_END_ROW}").Value = ""

                start = (p_idx-1)*STRUCTS_PER_PAGE
                end = min(start+STRUCTS_PER_PAGE, len(structs))
                page_structs = structs[start:end]

                row_base = LINE_START_ROW
                for struct_no, lamp_key, lamp_raw, led_w, cmt in page_structs:
                    installed = clean_watts(led_w)
                    excluded = str(struct_no).strip() in exclusions_set
                    exp_led = led_std_map.get(lamp_key, "")

                    # Removal description
                    desc_R = ""
                    is_led_suffix = bool(ENDSWITH_W_RE.match(str(lamp_raw or "")))
                    if is_led_suffix:
                        if exp_led:
                            desc_R = f"{clean_watts(lamp_raw)}W LED ST LT HEAD"
                        else:
                            m = ENDSWITH_W_RE.match(str(lamp_raw or ""))
                            watts = m.group(1) if m else ""
                            if watts:
                                desc_R = f"{clean_watts(lamp_raw)}W LED ST LT HEAD"
                    else:
                        if lamp_key in lpsv_map:
                            desc_R = f"{lpsv_map[lamp_key]}W LPSV ST LT HEAD"
                        elif lamp_key in mh_map:
                            desc_R = f"{mh_map[lamp_key]}W MH ST LT HEAD"
                        elif lamp_key in hpsv_map:
                            desc_R = f"{hpsv_map[lamp_key]}W HPSV ST LT HEAD"

                    # Install description
                    inst_for_desc = installed
                    desc_I = f"{inst_for_desc}W LED ST LT HEAD" if inst_for_desc else ""

                    # Write R row
                    ws.Range(f"{COL_STRUCTURE}{row_base}").Value = struct_no
                    ws.Range(f"{COL_LEGEND}{row_base}").Value = "R"
                    ws.Range(f"{COL_QTY_R}{row_base}").Value = 1
                    ws.Range(f"{COL_QTY_I}{row_base}").Value = ""
                    ws.Range(f"{COL_DESC}{row_base}").Value = desc_R

                    # Write I row
                    row_i = row_base + 1
                    ws.Range(f"{COL_STRUCTURE}{row_i}").Value = struct_no
                    ws.Range(f"{COL_LEGEND}{row_i}").Value = "I"
                    ws.Range(f"{COL_QTY_R}{row_i}").Value = ""
                    ws.Range(f"{COL_QTY_I}{row_i}").Value = 1
                    ws.Range(f"{COL_DESC}{row_i}").Value = desc_I

                    # Highlight if mismatch and NOT excluded
                    if (not excluded) and exp_led and installed and (installed != exp_led):
                        ws.Range(f"{COL_DESC}{row_base}").Interior.Color = 65535
                        ws.Range(f"{COL_DESC}{row_i}").Interior.Color = 65535

                    row_base += 2

            out_name = f"( JIS ) JOB INSTRUCTION SHEET {date_key}.xlsx"
            out_path = OUTPUT_DIR / out_name
            wb.SaveAs(str(out_path), 51)  # xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
            print(f"[JIS v{__version__}] Saved (COM): {out_path}")

        excel.Quit()
        return True
    except Exception as e:
        try:
            if excel: excel.Quit()
        except Exception:
            pass
        print("[JIS] COM mode failed:", e)
        return False

# ----------------- OPENPYXL OUTPUTS -----------------
def generate_outputs_openpyxl(groups, header_a2, af5_b2, ao35_c2, exclusions_set,
                              hpsv_map, mh_map, led_std_map, lpsv_map):
    YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")
    lvps = lpsv_map  # alias to guard against stray 'lvps' typos

    def as_excel_date(value):
        try:
            parsed = pd.to_datetime(str(value), errors="coerce")
            if pd.isna(parsed):
                return None, None
            return parsed.date(), "mm/dd/yyyy"
        except Exception:
            return None, None

    for date_key, g in groups.items():
        date_display = g["_date_display"].iloc[0]
        rows = g[["_Struct","_LampKey","_LampRaw","_LED_W","_Comments"]].values.tolist()
        pages = math.ceil(len(rows) / STRUCTS_PER_PAGE)

        wb = load_workbook(TEMPLATE_X)
        while len(wb.sheetnames) > 1:
            wb.remove(wb[wb.sheetnames[-1]])
        base = wb[wb.sheetnames[0]]
        base.title = "Sheet1"
        for i in range(2, pages+1):
            cp = wb.copy_worksheet(base); cp.title = f"Sheet{i}"
        sheets = [wb[f"Sheet{i}"] for i in range(1, pages+1)]

        # best-effort image copy
        imgs = getattr(base, "_images", [])
        for ws in sheets[1:]:
            for img in imgs:
                try:
                    path = getattr(img, "path", None)
                    if path:
                        new_img = XLImage(path)
                        new_img.anchor = img.anchor
                        try:
                            new_img.width, new_img.height = img.width, img.height
                        except Exception:
                            pass
                        ws.add_image(new_img)
                except Exception:
                    pass

        for p_idx, ws in enumerate(sheets, start=1):
            ws[CELL_DATE].value = date_display
            ws[CELL_HEADER].value = f"{header_a2} - VARIOUS LOCATIONS"
            ws[CELL_PAGE_CUR].value = p_idx
            ws[CELL_PAGE_TOT].value = pages
            ws[CELL_AF5].value = af5_b2
            ao35_date, fmt = as_excel_date(ao35_c2)
            ws[CELL_AO35].value = ao35_date if ao35_date else ""
            if ao35_date:
                ws[CELL_AO35].number_format = fmt

            # Clear line area
            for r in range(LINE_START_ROW, LINE_END_ROW+1):
                for col in (COL_STRUCTURE, COL_LEGEND, COL_QTY_I, COL_QTY_R, COL_DESC):
                    ws[f"{col}{r}"].value = ""
                    if col == COL_DESC:
                        ws[f"{col}{r}"].fill = PatternFill()

            start = (p_idx-1)*STRUCTS_PER_PAGE
            end = min(start+STRUCTS_PER_PAGE, len(rows))
            page_rows = rows[start:end]

            row_base = LINE_START_ROW
            for struct_no, lamp_key, lamp_raw, led_w, cmt in page_rows:
                installed = clean_watts(led_w)
                excluded = str(struct_no).strip() in exclusions_set
                exp_led = led_std_map.get(lamp_key, "")

                # Removal
                desc_R = ""
                is_led_suffix = bool(ENDSWITH_W_RE.match(str(lamp_raw or "")))
                if is_led_suffix:
                    if exp_led:
                        desc_R = f"{clean_watts(lamp_raw)}W LED ST LT HEAD"
                    else:
                        m = ENDSWITH_W_RE.match(str(lamp_raw or ""))
                        watts = m.group(1) if m else ""
                        if watts:
                            desc_R = f"{clean_watts(lamp_raw)}W LED ST LT HEAD"
                else:
                    if lamp_key in lpsv_map:
                        desc_R = f"{lpsv_map[lamp_key]}W LPSV ST LT HEAD"
                    elif lamp_key in mh_map:
                        desc_R = f"{mh_map[lamp_key]}W MH ST LT HEAD"
                    elif lamp_key in hpsv_map:
                        desc_R = f"{hpsv_map[lamp_key]}W HPSV ST LT HEAD"

                # Install
                inst_for_desc = installed
                desc_I = f"{inst_for_desc}W LED ST LT HEAD" if inst_for_desc else ""

                # Write rows
                ws[f"{COL_STRUCTURE}{row_base}"].value = struct_no
                ws[f"{COL_LEGEND}{row_base}"].value = "R"
                ws[f"{COL_QTY_R}{row_base}"].value = 1
                ws[f"{COL_QTY_I}{row_base}"].value = ""
                ws[f"{COL_DESC}{row_base}"].value = desc_R

                row_i = row_base + 1
                ws[f"{COL_STRUCTURE}{row_i}"].value = struct_no
                ws[f"{COL_LEGEND}{row_i}"].value = "I"
                ws[f"{COL_QTY_R}{row_i}"].value = ""
                ws[f"{COL_QTY_I}{row_i}"].value = 1
                ws[f"{COL_DESC}{row_i}"].value = desc_I

                # Highlight when mismatch and not excluded
                if (not excluded) and exp_led and installed and (installed != exp_led):
                    ws[f"{COL_DESC}{row_base}"].fill = YELLOW
                    ws[f"{COL_DESC}{row_i}"].fill = YELLOW

                row_base += 2

        out_name = f"( JIS ) JOB INSTRUCTION SHEET {date_key}.xlsx"
        out_path = OUTPUT_DIR / out_name
        wb.save(out_path)
        print(f"[JIS v{__version__}] Saved (openpyxl fallback): {out_path}")

# ----------------- main -----------------
def main():
    parser = argparse.ArgumentParser(description="JIS Page Populator v2.0")
    parser.add_argument("--assume-positional", action="store_true",
                        help="Skip header name warning/prompt and use positional columns A..E")
    args = parser.parse_args()

    print(f"[JIS v{__version__}] Starting...")
    for p in (INPUT_DIR, OUTPUT_DIR, CONV_DIR):
        if not p.exists():
            raise FileNotFoundError(f"Required folder/file not found: {p}")

    # Load inputs + conversion
    input_xlsx = find_first_xlsx(INPUT_DIR)
    header_a2, af5_b2, ao35_c2, exclusions_set = load_setup_fields_and_exclusions()
    hpsv_map, mh_map, led_std_map, lpsv_map = load_conversion_maps()
    raw_df = read_first_sheet_as_df(input_xlsx)
    groups, skipped_no_date_df = parse_input_and_group_by_date(raw_df, assume_positional=args.assume_positional)

    # --- Preflight log ---
    pre_rows = build_log_rows(groups, skipped_no_date_df, exclusions_set, hpsv_map, mh_map, led_std_map, lpsv_map)
    pre_path = write_excel_log(pre_rows, "Preflight")

    # Prompt user
    ans = input(f"Proceed to generate output files now? (Y/N): ").strip().lower()
    if ans not in ("y", "yes"):
        print("[JIS] Exiting without generating output files.")
        return

    # Re-read setup in case exclusions were updated
    header_a2, af5_b2, ao35_c2, exclusions_set2 = load_setup_fields_and_exclusions()

    # Generate outputs (COM first, fallback to openpyxl)
    ran = generate_outputs_com(groups, header_a2, af5_b2, ao35_c2, exclusions_set2,
                               hpsv_map, mh_map, led_std_map, lpsv_map)
    if not ran:
        print("[JIS] Falling back to openpyxl (images/controls may not copy).")
        generate_outputs_openpyxl(groups, header_a2, af5_b2, ao35_c2, exclusions_set2,
                                  hpsv_map, mh_map, led_std_map, lpsv_map)

    # --- Final log (rebuild with refreshed exclusions) ---
    final_rows = build_log_rows(groups, skipped_no_date_df, exclusions_set2, hpsv_map, mh_map, led_std_map, lpsv_map)
    write_excel_log(final_rows, "Final")

    print(f"[JIS v{__version__}] Done.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
