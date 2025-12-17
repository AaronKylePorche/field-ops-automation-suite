"""
Master JIS - Consolidate All JIS Output Files

This script reads all completed JIS output files from the output folder,
consolidates them into a single "Master JIS" workbook, and applies formatting.

Migrated to config-driven architecture.
Now reads all settings from config/config.py instead of hardcoded paths.
"""

import sys
from pathlib import Path
from datetime import datetime

# Import config from parent directories
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "config"))
import config

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# ---- Paths from config (now portable) ----
OUTPUT_DIR = Path(config.MASTERJIS_SETTINGS["jis_output_folder"])
SETUP_FILE = Path(config.MASTERJIS_SETTINGS["setup_file"])
MASTER_OUTPUT_DIR = Path(config.MASTERJIS_SETTINGS["master_output_folder"])

# Template geometry
LINE_START_ROW = 11
LINE_END_ROW = 32

def get_job_location():
    """Read Job Location from Setup.xlsx (cell A2) for filename."""
    try:
        setup_wb = load_workbook(SETUP_FILE, data_only=True)
        setup_ws = setup_wb.active
        job_location = str(setup_ws["A2"].value).strip().replace(" ", "_")
        return job_location
    except Exception as e:
        print(f"[WARNING] Could not read Setup.xlsx: {e}. Using default name.")
        return "JIS"

def main():
    print("[Master JIS] Starting consolidation...")

    # Ensure output directory exists
    MASTER_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Get job location for filename
    job_location = get_job_location()

    # Generate timestamped filename
    timestamp = datetime.now().strftime("%m.%d.%y_%H.%M")
    master_file = MASTER_OUTPUT_DIR / f"Master JIS {job_location} {timestamp}.xlsx"

    # Create master workbook
    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "Master"
    master_ws.append([
        "Line No.",
        "Reference/Structure Number",
        "Legend",
        "MATERIAL DESCRIPTION - LOCATION",
        "Install date",
        "JUNK"
    ])

    # Process each JIS output file
    jis_files = sorted(OUTPUT_DIR.glob("( JIS ) JOB INSTRUCTION SHEET *.xlsx"))

    if not jis_files:
        print(f"[Master JIS] No JIS output files found in {OUTPUT_DIR}")
        print("[Master JIS] Exiting without creating master file.")
        return

    print(f"[Master JIS] Found {len(jis_files)} JIS output file(s).")

    for file in jis_files:
        print(f'[Master JIS] Processing: {file.name}')
        try:
            wb = load_workbook(file, data_only=True)
            for sheet in wb.worksheets:
                print(f'  └─ Sheet: {sheet.title}')
                install_date = sheet["T3"].value

                for row in range(LINE_START_ROW, LINE_END_ROW + 1):
                    line_no = sheet[f"A{row}"].value
                    structure = sheet[f"B{row}"].value
                    legend = sheet[f"J{row}"].value
                    description = sheet[f"M{row}"].value

                    # Only process R (Removal) and I (Install) rows
                    if legend in ("I", "R"):
                        junk_value = 1 if legend == "R" else ""
                        new_row = [line_no, structure, legend, description, install_date, junk_value]
                        master_ws.append(new_row)

                        # Check merged cell M{row}:AC{row} for fill (highlighted rows)
                        merged_range = f"M{row}:AC{row}"
                        for merged_cell in sheet.merged_cells.ranges:
                            if str(merged_cell) == merged_range:
                                top_left_cell = sheet.cell(row=row, column=13)  # M column
                                fill = top_left_cell.fill
                                if fill.fill_type == "solid" and fill.fgColor and fill.fgColor.rgb:
                                    if fill.fgColor.rgb != "00000000":
                                        print(f'    [HIGHLIGHT] Row {row}: {fill.fgColor.rgb}')
                                        master_ws[f"D{master_ws.max_row}"].fill = PatternFill(
                                            fill_type="solid",
                                            fgColor=fill.fgColor.rgb
                                        )
        except Exception as e:
            print(f'[ERROR] Failed to process {file.name}: {e}')
            continue

    # Save master file
    master_wb.save(master_file)
    print(f'\n✅ [Master JIS] Master file created: {master_file}')

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
